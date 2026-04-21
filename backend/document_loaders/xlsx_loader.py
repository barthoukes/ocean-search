#!/usr/bin/env python3
"""
Loader for Excel files (XLSX, XLS).

This loader extracts data from Excel spreadsheets using openpyxl (for .xlsx)
or xlrd (for .xls). It converts each sheet to a text representation that
can be searched. Requires openpyxl or xlrd library.
"""

import os
import json
from typing import Optional, Set, Dict, Any, List
from langchain_core.documents import Document
from .base import DocumentLoader


class ExcelLoader(DocumentLoader):
    """Loader for Excel files (.xlsx, .xls)."""

    def _get_supported_extensions(self) -> Set[str]:
        return {'.xlsx', '.xls'}

    def load_document(self, file_path: str) -> Optional[Document]:
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"   ❌ File not found: {file_path}")
                return None
            
            # Check if file is readable
            if not os.access(file_path, os.R_OK):
                print(f"   ❌ Permission denied: Cannot read {file_path}")
                return None
            
            # Check file size
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                print(f"   ⚠️  File is empty: {file_path}")
                # Create a minimal document for empty file
                metadata = {
                    "filename": os.path.basename(file_path),
                    "filepath": file_path,
                    "type": "excel",
                    "extension": os.path.splitext(file_path)[1].lower(),
                    "size": 0,
                    "match_source": "metadata_only",
                    "embedding_type": "ollama",
                    "error": "empty_file",
                    "has_data": 0,
                    "total_sheets": 0,
                    "total_rows": 0,
                    "total_cells": 0,
                    "sheet_names": "",
                    "sheet_info": ""
                }
                content = f"Empty Excel file: {os.path.basename(file_path)}"
                return Document(page_content=content, metadata=metadata)
            
            # Try to load based on file extension
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.xlsx':
                return self._load_xlsx(file_path)
            elif ext == '.xls':
                return self._load_xls(file_path)
            else:
                print(f"   ❌ Unsupported Excel format: {ext}")
                return None
                
        except PermissionError:
            print(f"   ❌ Permission denied: Cannot access {file_path}")
            return None
        except OSError as e:
            print(f"   ❌ OS error loading {file_path}: {e}")
            return None
        except Exception as e:
            print(f"   ❌ Unexpected error loading Excel file {file_path}: {e}")
            return None

    def _load_xlsx(self, file_path: str) -> Optional[Document]:
        """Load .xlsx files using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            print(f"   ❌ openpyxl not installed. Please install: pip install openpyxl")
            print(f"   💡 Tip: For .xlsx files, run: pip install openpyxl")
            return None

        try:
            # Suppress the DrawingML warning
            import warnings
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")
            
            print(f"   📊 Loading Excel file: {os.path.basename(file_path)}")
            
            # Try to load with different options if first attempt fails
            workbook = None
            try:
                workbook = load_workbook(file_path, data_only=True)
            except Exception as e1:
                print(f"   ⚠️  Failed to load with data_only=True: {e1}")
                try:
                    print(f"   🔄 Trying with data_only=False...")
                    workbook = load_workbook(file_path, data_only=False)
                except Exception as e2:
                    print(f"   ⚠️  Failed to load with data_only=False: {e2}")
                    return None
            
            if workbook is None:
                print(f"   ❌ Could not load workbook: {file_path}")
                return None
            
            content_parts = []
            
            # Store sheet info as JSON string instead of list of dicts
            sheet_names = []
            sheet_info_str = []
            total_rows = 0
            total_cells = 0
            sheets_with_data = 0
            
            print(f"   📑 Found {len(workbook.sheetnames)} sheet(s)")
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get sheet dimensions
                max_row = sheet.max_row or 0
                max_col = sheet.max_column or 0
                
                sheet_names.append(sheet_name)
                sheet_info_str.append(f"{sheet_name}:{max_row}x{max_col}")
                total_rows += max_row
                
                # Build content for this sheet
                sheet_content = f"\n[Sheet: {sheet_name}]\n"
                sheet_content += f"Dimensions: {max_row} rows × {max_col} columns\n\n"
                
                # Extract data from rows
                row_count = 0
                error_count = 0
                
                try:
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        try:
                            # Filter out completely empty rows
                            non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
                            if not non_empty:
                                continue
                            
                            row_count += 1
                            
                            # Format row data as text
                            row_texts = []
                            for col_idx, cell in enumerate(row):
                                if cell is not None and str(cell).strip():
                                    try:
                                        # Convert column index to letter (A, B, C, etc.)
                                        col_letter = self._col_index_to_letter(col_idx)
                                        cell_value = str(cell).strip()
                                        row_texts.append(f"{col_letter}{row_idx}: {cell_value}")
                                        total_cells += 1
                                    except Exception as e:
                                        error_count += 1
                                        if error_count <= 3:  # Limit error messages
                                            print(f"      ⚠️  Error formatting cell {col_idx+1},{row_idx}: {e}")
                            
                            if row_texts:
                                sheet_content += f"Row {row_idx}: " + " | ".join(row_texts) + "\n"
                                
                        except Exception as e:
                            error_count += 1
                            if error_count <= 3:
                                print(f"      ⚠️  Error processing row {row_idx}: {e}")
                            continue
                    
                    if row_count > 0:
                        content_parts.append(sheet_content)
                        sheets_with_data += 1
                        print(f"      ✓ Sheet '{sheet_name}': {row_count} rows with data")
                    else:
                        content_parts.append(f"\n[Sheet: {sheet_name}] - Empty sheet\n")
                        print(f"      ℹ️  Sheet '{sheet_name}': No data found")
                        
                    if error_count > 3:
                        print(f"      ⚠️  {error_count} total errors in sheet '{sheet_name}'")
                        
                except Exception as e:
                    print(f"      ❌ Error iterating through sheet '{sheet_name}': {e}")
                    content_parts.append(f"\n[Sheet: {sheet_name}] - Error reading sheet: {e}\n")
                    continue
            
            # Combine all sheet contents
            if content_parts:
                content = "\n".join(content_parts)
            else:
                content = f"Excel file with {len(workbook.sheetnames)} sheets - no readable data found"
                print(f"   ⚠️  No readable data extracted from any sheet")
            
            # Create flat metadata (no nested structures)
            metadata = {
                "filename": os.path.basename(file_path),
                "filepath": file_path,
                "type": "excel",
                "extension": ".xlsx",
                "size": os.path.getsize(file_path),
                "match_source": "file_content",
                "embedding_type": "ollama",
                "total_sheets": len(workbook.sheetnames),
                "total_rows": total_rows,
                "total_cells": total_cells,
                "sheets_with_data": sheets_with_data,
                "has_data": 1 if total_cells > 0 else 0,
                "sheet_names": ",".join(sheet_names),
                "sheet_info": "; ".join(sheet_info_str)
            }
            
            if total_cells == 0:
                print(f"   ⚠️  No data cells found in {os.path.basename(file_path)}")
            else:
                print(f"   ✅ Successfully loaded: {total_cells} cells from {sheets_with_data}/{len(workbook.sheetnames)} sheets")
            
            return Document(page_content=content, metadata=metadata)
            
        except Exception as e:
            print(f"   ❌ Error loading XLSX file {file_path}: {e}")
            print(f"   💡 Tip: Try opening the file in Excel and resaving it")
            return None

    def _load_xls(self, file_path: str) -> Optional[Document]:
        """Load .xls files using xlrd (older Excel format)."""
        try:
            import xlrd
        except ImportError:
            print(f"   ❌ xlrd not installed. Please install: pip install xlrd")
            print(f"   💡 Tip: For .xls files, run: pip install xlrd")
            return None

        try:
            print(f"   📊 Loading legacy Excel file: {os.path.basename(file_path)}")
            
            # xlrd 2.0+ doesn't support .xlsx, only .xls
            try:
                workbook = xlrd.open_workbook(file_path)
            except xlrd.XLRDError as e:
                if "Excel xlsx file" in str(e):
                    print(f"   ❌ This appears to be an .xlsx file with .xls extension")
                    print(f"   💡 Tip: Rename the file to .xlsx or use openpyxl")
                else:
                    print(f"   ❌ xlrd error: {e}")
                return None
            
            content_parts = []
            
            # Store sheet info as flat strings
            sheet_names = []
            sheet_info_str = []
            total_rows = 0
            total_cells = 0
            sheets_with_data = 0
            
            print(f"   📑 Found {workbook.nsheets} sheet(s)")
            
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                
                sheet_names.append(sheet_name)
                sheet_info_str.append(f"{sheet_name}:{sheet.nrows}x{sheet.ncols}")
                total_rows += sheet.nrows
                
                # Build content for this sheet
                sheet_content = f"\n[Sheet: {sheet_name}]\n"
                sheet_content += f"Dimensions: {sheet.nrows} rows × {sheet.ncols} columns\n\n"
                
                # Extract data from rows
                row_count = 0
                error_count = 0
                
                for row_idx in range(sheet.nrows):
                    try:
                        row = sheet.row_values(row_idx)
                        
                        # Filter out completely empty rows
                        non_empty = [str(cell).strip() for cell in row if cell and str(cell).strip()]
                        if not non_empty:
                            continue
                        
                        row_count += 1
                        
                        # Format row data as text
                        row_texts = []
                        for col_idx, cell in enumerate(row):
                            if cell is not None and str(cell).strip():
                                try:
                                    col_letter = self._col_index_to_letter(col_idx)
                                    cell_value = str(cell).strip()
                                    row_texts.append(f"{col_letter}{row_idx + 1}: {cell_value}")
                                    total_cells += 1
                                except Exception as e:
                                    error_count += 1
                                    if error_count <= 3:
                                        print(f"      ⚠️  Error formatting cell {col_idx+1},{row_idx+1}: {e}")
                        
                        if row_texts:
                            sheet_content += f"Row {row_idx + 1}: " + " | ".join(row_texts) + "\n"
                            
                    except Exception as e:
                        error_count += 1
                        if error_count <= 3:
                            print(f"      ⚠️  Error processing row {row_idx + 1}: {e}")
                        continue
                
                if row_count > 0:
                    content_parts.append(sheet_content)
                    sheets_with_data += 1
                    print(f"      ✓ Sheet '{sheet_name}': {row_count} rows with data")
                else:
                    content_parts.append(f"\n[Sheet: {sheet_name}] - Empty sheet\n")
                    print(f"      ℹ️  Sheet '{sheet_name}': No data found")
                    
                if error_count > 3:
                    print(f"      ⚠️  {error_count} total errors in sheet '{sheet_name}'")
            
            # Combine all sheet contents
            if content_parts:
                content = "\n".join(content_parts)
            else:
                content = f"Excel file with {workbook.nsheets} sheets - no readable data found"
                print(f"   ⚠️  No readable data extracted from any sheet")
            
            # Create flat metadata (no nested structures)
            metadata = {
                "filename": os.path.basename(file_path),
                "filepath": file_path,
                "type": "excel",
                "extension": ".xls",
                "size": os.path.getsize(file_path),
                "match_source": "file_content",
                "embedding_type": "ollama",
                "total_sheets": workbook.nsheets,
                "total_rows": total_rows,
                "total_cells": total_cells,
                "sheets_with_data": sheets_with_data,
                "has_data": 1 if total_cells > 0 else 0,
                "sheet_names": ",".join(sheet_names),
                "sheet_info": "; ".join(sheet_info_str)
            }
            
            if total_cells == 0:
                print(f"   ⚠️  No data cells found in {os.path.basename(file_path)}")
            else:
                print(f"   ✅ Successfully loaded: {total_cells} cells from {sheets_with_data}/{workbook.nsheets} sheets")
            
            return Document(page_content=content, metadata=metadata)
            
        except Exception as e:
            print(f"   ❌ Error loading XLS file {file_path}: {e}")
            print(f"   💡 Tip: Try converting the file to .xlsx format")
            return None

    def _col_index_to_letter(self, col_idx: int) -> str:
        """Convert column index (0-based) to Excel column letter (A, B, C, ..., Z, AA, AB, etc.)"""
        if col_idx < 0:
            return "?"
        result = ""
        n = col_idx
        while n >= 0:
            n, remainder = divmod(n, 26)
            result = chr(65 + remainder) + result
            n -= 1
        return result

    def extract_tables(self, file_path: str) -> Optional[str]:
        """
        Optional method to extract structured tables from Excel files.
        Returns JSON string instead of list of dicts for ChromaDB compatibility.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("   ❌ openpyxl not installed, cannot extract tables")
            print("   💡 Tip: pip install openpyxl")
            return None

        try:
            if not os.path.exists(file_path):
                print(f"   ❌ File not found: {file_path}")
                return None
                
            workbook = load_workbook(file_path, data_only=True)
            tables_data = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Try to find if there are named tables
                if hasattr(sheet, 'tables') and sheet.tables:
                    for table_name, table in sheet.tables.items():
                        table_data = {
                            "sheet": sheet_name,
                            "name": table_name,
                            "ref": table.ref,
                            "row_count": 0
                        }
                        
                        # Extract table data as text
                        cell_range = table.ref.split(':')
                        if len(cell_range) == 2:
                            start_cell, end_cell = cell_range
                            # Extract column letters and row numbers
                            import re
                            start_col = re.match(r'([A-Z]+)', start_cell).group(1)
                            start_row = int(re.search(r'(\d+)', start_cell).group(1))
                            end_col = re.match(r'([A-Z]+)', end_cell).group(1)
                            end_row = int(re.search(r'(\d+)', end_cell).group(1))
                            
                            # Get headers (first row of table)
                            headers = []
                            for col_idx in range(len(start_col)):
                                cell = sheet[f"{chr(ord(start_col) + col_idx)}{start_row}"]
                                if cell.value:
                                    headers.append(str(cell.value))
                            
                            # Count data rows
                            table_data["row_count"] = max(0, end_row - start_row)
                            table_data["headers"] = ",".join(headers)
                        
                        tables_data.append(table_data)
            
            # Return as JSON string if tables found
            if tables_data:
                print(f"   📊 Found {len(tables_data)} named tables in {file_path}")
                return json.dumps(tables_data)
            else:
                print(f"   ℹ️  No named tables found in {file_path}")
                return None
            
        except ImportError:
            print(f"   ❌ openpyxl not installed")
            return None
        except Exception as e:
            print(f"   ❌ Error extracting tables from {file_path}: {e}")
            return None


if __name__ == "__main__":
    # Self-test: check if libraries are installed and create test file
    try:
        from openpyxl import Workbook
        openpyxl_available = True
    except ImportError:
        openpyxl_available = False

    if not openpyxl_available:
        print("ExcelLoader self-test skipped: openpyxl not installed")
        print("Install with: pip install openpyxl")
        print("For .xls support also: pip install xlrd")
    else:
        print("ExcelLoader self-test:")
        loader = ExcelLoader()
        
        # Test can_handle
        print(f"  can_handle('test.xlsx'): {loader.can_handle('test.xlsx')}")
        print(f"  can_handle('test.xls'): {loader.can_handle('test.xls')}")
        print(f"  can_handle('test.txt'): {loader.can_handle('test.txt')}")
        
        # Create a test Excel file
        import tempfile
        
        wb = Workbook()
        
        # Sheet 1: Sample data
        ws1 = wb.active
        ws1.title = "Employees"
        ws1['A1'] = "Name"
        ws1['B1'] = "Department"
        ws1['C1'] = "Salary"
        ws1['A2'] = "John Doe"
        ws1['B2'] = "Engineering"
        ws1['C2'] = 75000
        ws1['A3'] = "Jane Smith"
        ws1['B3'] = "Marketing"
        ws1['C3'] = 68000
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Load the test file
            doc = loader.load_document(tmp_path)
            if doc:
                print("  Load on test Excel file SUCCESS")
                print(f"    Content length: {len(doc.page_content)} characters")
                print(f"    Metadata: {doc.metadata}")
                print(f"\n  Content preview:")
                print("    " + doc.page_content[:300].replace('\n', '\n    '))
            else:
                print("  Load on test Excel file FAILED")
        finally:
            os.unlink(tmp_path)
            
        # Test with non-existent file
        print("\n  Testing error handling:")
        doc = loader.load_document("non_existent_file.xlsx")
        if doc is None:
            print("  ✓ Non-existent file correctly returned None")

            